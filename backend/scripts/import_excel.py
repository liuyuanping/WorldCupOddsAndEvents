"""Import all data from Excel to offline database."""
import json
import openpyxl
import httpx
import asyncio

API = "http://localhost:8000"

# Map Chinese team names → Polymarket team_ids
TEAM_MAP = {
    "墨西哥": "mexico", "南非": "south_africa", "韩国": "south_korea", "捷克": "czechia",
    "加拿大": "canada", "波黑": "bosnia", "卡塔尔": "qatar", "瑞士": "switzerland",
    "巴西": "brazil", "摩洛哥": "morocco", "海地": "haiti", "苏格兰": "scotland",
    "美国": "usa", "巴拉圭": "paraguay", "澳大利亚": "australia", "土耳其": "turkiye",
    "德国": "germany", "库拉索": "curacao", "科特迪瓦": "ivory_coast", "厄瓜多尔": "ecuador",
    "荷兰": "netherlands", "日本": "japan", "瑞典": "sweden", "突尼斯": "tunisia",
    "比利时": "belgium", "埃及": "egypt", "伊朗": "iran", "新西兰": "new_zealand",
    "西班牙": "spain", "佛得角": "cape_verde", "沙特": "saudi_arabia", "乌拉圭": "uruguay",
    "法国": "france", "塞内加尔": "senegal", "伊拉克": "iraq", "挪威": "norway",
    "阿根廷": "argentina", "阿尔及利亚": "algeria", "奥地利": "austria", "约旦": "jordan",
    "葡萄牙": "portugal", "刚果(金)": "congo_dr", "乌兹别克斯坦": "uzbekistan", "哥伦比亚": "colombia",
    "英格兰": "england", "克罗地亚": "croatia", "加纳": "ghana", "巴拿马": "panama",
    "刚果金": "congo_dr", "沙特阿拉伯": "saudi_arabia",
}

# Severity mapping
SEV_MAP = {"小": 1, "中等": 2, "较大": 3, "大": 4, "正面": 2}
SEV_MAP2 = {"小": 1, "中等": 2, "中等": 2, "大": 3, "大": 4, "正面": 2}

async def clear_existing():
    async with httpx.AsyncClient() as c:
        # Clear all database events
        r = await c.get(f"{API}/api/v1/champion/events?provider=database")
        events = r.json().get("events", [])
        deleted = 0
        for e in events:
            sid = e.get("source_id", "")
            if sid:
                await c.delete(f"{API}/api/v1/champion/events/db/{sid}")
                deleted += 1
        print(f"Cleared {deleted} existing database events")

async def add_event(team_id, team_name, evt_type, title, desc, timestamp, severity, confidence=1.0, source_url=""):
    async with httpx.AsyncClient() as c:
        r = await c.post(f"{API}/api/v1/champion/events/db", json={
            "team_id": team_id, "team_name": team_name, "event_type": evt_type,
            "title": title, "description": desc, "timestamp": timestamp,
            "severity": severity, "confidence": confidence, "source_url": source_url,
        })
        if r.status_code != 200:
            print(f"  FAILED {r.status_code}: {team_name} - {title[:40]}")

async def import_matches():
    wb = openpyxl.load_workbook('/app/data/2026美加墨世界杯数据汇总.xlsx')
    ws = wb['比赛详细信息']
    count = 0
    for r in range(2, ws.max_row + 1):
        group = ws.cell(r, 1).value
        round_label = ws.cell(r, 2).value
        home = ws.cell(r, 3).value
        away = ws.cell(r, 4).value
        score = ws.cell(r, 5).value
        dt_str = ws.cell(r, 6).value
        venue = ws.cell(r, 7).value or ""
        goals = ws.cell(r, 8).value or ""
        result = ws.cell(r, 9).value or ""
        notes = ws.cell(r, 10).value or ""

        if not all([group, home, away, score, dt_str]):
            continue

        # Parse score
        if ":" in score:
            hs, aws = score.split(":")
        elif "-" in score:
            hs, aws = score.split("-")
        else:
            continue

        # Parse datetime - convert Chinese format to ISO
        ts = dt_str.replace(" ", "T") + ":00Z"
        if "T" not in ts:
            ts = f"{dt_str}T00:00:00Z"

        home_id = TEAM_MAP.get(home)
        away_id = TEAM_MAP.get(away)
        if not home_id or not away_id:
            print(f"  SKIP unknown team: {home}({home_id}) vs {away}({away_id})")
            continue

        desc = f"2026世界杯{group} · {home} {score} {away} · {venue}"
        if notes and notes != "None":
            desc += f" · {notes}"
        if goals and goals != "None":
            desc += f"\n进球: {goals}"

        # Home team event
        await add_event(home_id, home, "match_result",
            f"{home} {score} {away} (vs {away})",
            desc, ts, 2, 1.0)
        count += 1

        # Away team event
        score_r = f"{aws}-{hs}"
        await add_event(away_id, away, "match_result",
            f"{away} {score_r} {home} (vs {home})",
            desc, ts, 2, 1.0)
        count += 1

    print(f"Imported {count} match events ({count//2} matches)")

async def import_events():
    wb = openpyxl.load_workbook('/app/data/2026美加墨世界杯数据汇总.xlsx')
    ws = wb['球队重大变故']
    count = 0
    for r in range(2, ws.max_row + 1):
        team_cn = ws.cell(r, 1).value
        etype = ws.cell(r, 2).value or "other"
        player = ws.cell(r, 3).value or ""
        detail = ws.cell(r, 4).value or ""
        impact = ws.cell(r, 5).value or ""
        time_desc = ws.cell(r, 6).value or ""
        champion_impact = ws.cell(r, 7).value or ""

        if not team_cn:
            continue

        team_id = TEAM_MAP.get(team_cn)
        if not team_id:
            print(f"  SKIP unknown team: {team_cn}")
            continue

        # Map event type
        evt_type = "other"
        type_lower = etype.lower() if etype else ""
        if "伤病" in type_lower or "受伤" in type_lower:
            evt_type = "injury"
        elif "状态" in type_lower or "里程碑" in type_lower:
            evt_type = "form"
        elif "天气" in type_lower:
            evt_type = "weather"
        elif "内讧" in type_lower or "矛盾" in type_lower:
            evt_type = "conflict"
        elif "红牌" in type_lower or "停赛" in type_lower:
            evt_type = "suspension"
        elif "纪录" in type_lower or "突破" in type_lower or "最快进球" in type_lower:
            evt_type = "record"
        elif "爆冷" in type_lower or "惨败" in type_lower:
            evt_type = "shock"
        elif "出局" in type_lower:
            evt_type = "elimination"

        # Map severity
        severity = 2  # default
        if impact:
            impact_lower = str(impact).lower()
            if "大" in impact_lower and "较" not in impact_lower and "正" not in impact_lower:
                severity = 4
            elif "较大" in impact_lower or "大" in impact_lower:
                severity = 3
            elif "中等" in impact_lower or "中" in impact_lower:
                severity = 2
            elif "小" in impact_lower:
                severity = 1
            elif "正" in impact_lower:  # 正面
                severity = 2

        # Timestamp — use best guess
        ts = "2026-06-24T12:00:00Z"
        if time_desc:
            if "第1轮" in time_desc:
                ts = "2026-06-18T12:00:00Z"
            elif "第2轮" in time_desc:
                ts = "2026-06-24T12:00:00Z"
            elif "开赛前" in time_desc:
                ts = "2026-06-11T08:00:00Z"
            elif "小组赛" in time_desc:
                ts = "2026-06-16T12:00:00Z"

        title = f"{player} - {etype}" if player else etype
        if etype == "伤病":
            title = f"{player}受伤"
        elif etype == "状态" or etype == "状态火热" or etype == "状态回升":
            title = f"{player}状态出色"
        elif etype == "纪录/状态" or etype == "纪录":
            title = f"{player}创造纪录"

        desc = detail
        if champion_impact and champion_impact != "None":
            desc += f"\n影响: {champion_impact}"

        await add_event(team_id, team_cn, evt_type, title, desc, ts, severity, 0.85)
        count += 1

    print(f"Imported {count} team events")

async def main():
    print("=== Step 1: Clear existing ===")
    await clear_existing()
    print("\n=== Step 2: Import matches ===")
    await import_matches()
    print("\n=== Step 3: Import team events ===")
    await import_events()
    print("\n=== Step 4: Verify ===")
    async with httpx.AsyncClient() as c:
        r = await c.get(f"{API}/api/v1/champion/events?provider=database&limit=10")
        data = r.json()
        print(f"Total events: {data['total']}")
        for e in data['events'][:8]:
            print(f"  [{e['severity']}] {e['team_name']}: {e['title'][:60]}")

asyncio.run(main())
